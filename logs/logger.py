from datetime import  timedelta
from openpyxl import Workbook, load_workbook
from plistlib import InvalidFileException
import zipfile
from datetime import datetime

INTERVAL_MINUTES = 10  # Интервал записи (10 минут)

class Logger:
    def getNextLogTime(self, current_time):
        next_time = current_time.replace(second=0, microsecond=0) + \
                    timedelta(minutes=INTERVAL_MINUTES - (current_time.minute % INTERVAL_MINUTES))
        return next_time
    
    def saveToExcel(self,pair, event, teeth, angle, comment, fileName): 
        try:
            # Пытаемся загрузить существующий файл
            workbook = load_workbook(fileName)
            sheet = workbook.active
        #except FileNotFoundError:
            # Если файла нет — создаем новый
            #workbook = Workbook()
            #sheet = workbook.active
            #sheet.append(["Дата", "Событие", "Пара", "Зубы (Teeth)", "Угол", "Комментарий"])
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException) as e:
            print(f"⚠️ Ошибка при загрузке файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return
        
        # Добавляем новую строку
        sheet.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            event,
            pair,
            teeth,
            angle,
            comment
        ])
        
        try:
            workbook.save(fileName)
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException, PermissionError) as e:
            print(f"⚠️ Ошибка при сохранении файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return
    
    def saveToExcelOpenAi(self,pair, event, positionType, price, sl, tp, signalPower, trendAnalysis, volumeAnalysis, summary, error, fileName) : 
        try:
            # Пытаемся загрузить существующий файл
            workbook = load_workbook(fileName)
            sheet = workbook.active
        except FileNotFoundError:
            # Если файла нет — создаем новый
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Дата", "Событие", "Пара", "Тип ордера", "Цена", "sl", "tp", "мощность сигнала", "Анализ тренда", "Анализ объема", "Итого", "Ошибка"])
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException) as e:
            print(f"⚠️ Ошибка при загрузке файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return
        
        # Добавляем новую строку
        sheet.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            event,
            pair,
            positionType, 
            price, 
            sl, 
            tp, 
            signalPower, 
            trendAnalysis, 
            volumeAnalysis, 
            summary, 
            error
        ])
        
        try:
            workbook.save(fileName)
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException, PermissionError) as e:
            print(f"⚠️ Ошибка при сохранении файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return
  
    def saveErrorsToExcel(self, service, comment, fileName): 
        try:
            # Пытаемся загрузить существующий файл
            workbook = load_workbook(fileName)
            sheet = workbook.active

        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException) as e:
            
            print(f"⚠️ Ошибка при загрузке файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Дата", "сервис", "ошибка"])
            
            return
        
        # Добавляем новую строку
        sheet.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            service,
            comment
        ])
        
        try:
            workbook.save(fileName)
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException, PermissionError) as e:
            print(f"⚠️ Ошибка при сохранении файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return

    def saveBBToExcel(self, pair, typeOperation, price, lower, middle, upper, comment, fileName): 
        try:
            # Пытаемся загрузить существующий файл
            workbook = load_workbook(fileName)
            sheet = workbook.active

        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException) as e:
            
            print(f"⚠️ Ошибка при загрузке файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["время", "пара", "тип операции", "цена", "lower", "middle", "upper", "комментарий"])
            
            return
        
        # Добавляем новую строку
        sheet.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            pair,
            typeOperation,
            price,
            lower,
            middle,
            upper,
            comment           
            
        ])
        
        try:
            workbook.save(fileName)
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException, PermissionError) as e:
            print(f"⚠️ Ошибка при сохранении файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return
