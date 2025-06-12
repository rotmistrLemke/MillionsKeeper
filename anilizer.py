from plistlib import InvalidFileException
import zipfile
import numpy as np
from appEnum import IndicatorType,TargetType,Settings
import math
from decimal import Decimal
import pandas as pd
import MetaTrader5 as mt5
from openpyxl import Workbook, load_workbook
from datetime import datetime


class ZeroIntersection:    
    def check(self, cciValues):
        if len(cciValues) < 2:
            return {"value": False}
                    
        currentValue = cciValues[0]
        previousValue = cciValues[1]
        
        if currentValue > 0 and currentValue < 15 and previousValue < -10:
            return {"value": True, "target": TargetType.LONG}
        
        if currentValue < 0 and currentValue > -15 and previousValue > 10:
            return {"value": True, "target": TargetType.SHORT}
            
        return {"value": False}

class HundredIntersection:
    def check(self, cciValues):
        if len(cciValues) < 2:
            return {"value": False}
            
        currentValue = cciValues[0]
        previousValue = cciValues[1]
        
        if currentValue > 100 and currentValue < 115 and previousValue < 90:
            return {"value": True, "target": TargetType.LONG}
        
        if currentValue < -100 and currentValue > -115 and previousValue > -90:
            return {"value": True, "target": TargetType.SHORT}
            
        return {"value": False}

class Extremum:    
    def __init__(self,settings):
        self.CCI_ReferenceLimitForEnter = settings["CCI_ReferenceLimitForEnter"]
    
    def tryAngleCoefficient(self,current,prev):
        current = abs(current)
        prev = abs(prev)
        if current < prev:
            return (prev-current)/prev
        elif current > prev:
            return (current-prev)/current
    
    def angleForCciStoch(self,currentValue, previousValue, y, degrees=True):
            """
            Вычисляет arctg(x) и возвращает угол в градусах или радианах.

            Параметры:
                x (float): Число, для которого вычисляется арктангенс.
                degrees (bool): Если True, возвращает угол в градусах, иначе в радианах.

            Возвращает:
                float: Угол в градусах или радианах.
            """
            x = (currentValue - previousValue)
            angle_rad = math.atan2(x, y)
            return int(f"{math.degrees(angle_rad):.0f}") if degrees else int(f"{angle_rad:.0f}")

    def cciReverse(self, cciValues, limit):
        if len(cciValues) < 2:
            return {"value": False}
            
        currentValue = cciValues[0]
        previousValue = cciValues[1]
        # coefficient = self.tryAngleCoefficient(currentValue,previousValue)
        angle = self.angleForCciStoch(currentValue,previousValue,50)
        
        if currentValue > limit and previousValue > currentValue and angle < -20:
            return {"value": True, "target": TargetType.SHORT, "angle": angle}
        
        if currentValue < limit * -1 and previousValue < currentValue and angle > 20:
            return {"value": True, "target": TargetType.LONG, "angle": angle}
            
        return {"value": False, "angle": angle}

    def stochasticReverse(self, stochasticValues):
        if len(stochasticValues) < 2:
            return {"value": False}
            
        currentValue = stochasticValues[0]
        previousValue = stochasticValues[1]
        # coefficient = self.tryAngleCoefficient(currentValue,previousValue)
        angle = self.angleForCciStoch(currentValue,previousValue,10)

        if previousValue > currentValue and angle < -15:
            return {"value": True, "target": TargetType.SHORT, "angle":angle }
        
        if previousValue < currentValue and angle > 15:
            return {"value": True, "target": TargetType.LONG, "angle":angle}
            
        return {"value": False, "angle": angle}

    def checkForEnter(self, cciValues, stochasticValues):
        """Основной метод проверки условий"""
        cciReverse_result = self.cciReverse(cciValues, self.CCI_ReferenceLimitForEnter)
        stochasticReverse_result = self.stochasticReverse(stochasticValues)
        
        if cciReverse_result["value"] and stochasticReverse_result["value"]:
            if (cciReverse_result["target"] == TargetType.LONG and 
                stochasticReverse_result["target"] == TargetType.LONG):
                return {"value": True, "target": TargetType.LONG, "cciAngle": cciReverse_result["angle"], "stochAngle": stochasticReverse_result["angle"]}
            
            if (cciReverse_result["target"] == TargetType.SHORT and 
                stochasticReverse_result["target"] == TargetType.SHORT):
                return {"value": True, "target": TargetType.SHORT, "cciAngle": cciReverse_result["angle"], "stochAngle": stochasticReverse_result["angle"]}
        
        return {"value": False, "cciAngle": cciReverse_result["angle"], "stochAngle": stochasticReverse_result["angle"]}
    
    def checkForClose(self, cciValues):
        """Основной метод проверки условий"""
        currentValue = cciValues[0]
        
        if currentValue > -20:
            return {"value": True, "cciValue": currentValue}
        
        if currentValue < 20:
            return {"value": True, "cciValue": currentValue}
            
        return {"value": False, "cciValue": currentValue}
    
class Alligator:

    def smma(self,data, period):
        smma_values = []
        for i in range(len(data)):
            if i < period:
                smma_values.append(np.nan)
            elif i == period:
                smma_values.append(data[i-period:i].mean())
            else:
                smma_values.append((smma_values[-1] * (period - 1) + data[i]) / period)
        return pd.Series(smma_values)
    
    def angle(self,currentLipsValue, previousLipsValue, pair, pairXvalue, degrees=True):
        """
        Вычисляет arctg(x) и возвращает угол в градусах или радианах.

        Параметры:
            x (float): Число, для которого вычисляется арктангенс.
            degrees (bool): Если True, возвращает угол в градусах, иначе в радианах.

        Возвращает:
            float: Угол в градусах или радианах.
        """
        x = (currentLipsValue - previousLipsValue) / mt5.symbol_info(pair).point
        angle_rad = math.atan2(x, pairXvalue/2)
        return int(f"{math.degrees(angle_rad):.0f}") if degrees else int(f"{angle_rad:.0f}")
    
    def CountDecimalPlace(self,pair):    
        num = Decimal(str(mt5.symbol_info(pair).point))
        return  abs(num.as_tuple().exponent)
    
    def getAlligatorVsCurrentCandelDiff(self,pair, alligatorValue):
        """Возвращает разницу между текущей ценой и индикатором аллигатор по модулю."""
        return int(f"{abs(alligatorValue - mt5.symbol_info_tick(pair).bid)/ mt5.symbol_info(pair).point:.0f}")
    
    def getLipsVsTeethDiff(self,pair, lips, teeth):
        """Возвращает разницу между текущей ценой и индикатором аллигатор по модулю."""
        return abs(lips - teeth)/ mt5.symbol_info(pair).point
    
    def saveToExcel(self,pair, event, teeth, angle, comment, fileName): 
        try:
            # Пытаемся загрузить существующий файл
            workbook = load_workbook(fileName)
            sheet = workbook.active
        #except FileNotFoundError:
            # Если файла нет — создаем новый
            #workbook = Workbook()
            #sheet = workbook.active
            #sheet.append(["Дата", "Событие", "Пара", "Зубы (Teeth)", "Угол", "Комментарий"])
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException) as e:
            print(f"⚠️ Ошибка при загрузке файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return
        
        # Добавляем новую строку
        sheet.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            event,
            pair,
            teeth,
            angle,
            comment
        ])
        
        try:
            workbook.save(fileName)
        except (FileNotFoundError, zipfile.BadZipFile, InvalidFileException, PermissionError) as e:
            print(f"⚠️ Ошибка при сохранении файла {fileName}: {e}")
            print("Продолжаю работу без сохранения в Excel...")
            return
        
    
    def MainData(self,df):
        medianPrice = (df['high'] + df['low']) / 2  # Медианная цена (HL/2)
        openPrice = df['open'].iloc[-1]
        # Рассчитываем линии Аллигатора
        jaw = self.smma(medianPrice, 13)  # Челюсти (13)
        teeth = self.smma(medianPrice, 8)   # Зубы (8)
        lips = self.smma(medianPrice, 5)    # Губы (5)
        return medianPrice,jaw,teeth,lips,openPrice

    def ShiftedData(self,jaw,teeth,lips,medianPrice):
        # Рассчитываем линии Аллигатора
        jaw = self.smma(medianPrice, 13)  # Челюсти (13)
        teeth = self.smma(medianPrice, 8)   # Зубы (8)
        lips = self.smma(medianPrice, 5)    # Губы (5)
        # Смещаем линии  (бары 3, 1, -1)
        jawShifted = jaw.shift(3)
        teethShifted = teeth.shift(1)
        lipsShifted = lips.shift(-1)

        return jawShifted,teethShifted,lipsShifted
    
    def LastData(self,pair,jawShifted,teethShifted,lipsShifted): 
        countDecimalPlace = self.CountDecimalPlace(pair)
        lastJaw = float(f"{jawShifted.iloc[-2]:.{countDecimalPlace}f}")
        lastTeeth =  float(f"{teethShifted.iloc[-2]:.{countDecimalPlace}f}")
        lastLips = float(f"{lipsShifted.iloc[-2]:.{countDecimalPlace}f}")
        prelastLips = float(f"{lipsShifted.iloc[-3]:.{countDecimalPlace}f}")
        return lastJaw,lastTeeth,lastLips,prelastLips
    
    def SupportData(self,lastLips,prelastLips,pair,dictPairXvalue, lastTeeth):
        angle = self.angle(lastLips,prelastLips,pair,dictPairXvalue.get(pair, 100))
        candleDiff = self.getAlligatorVsCurrentCandelDiff(pair,lastLips)
        lipsVsTeethDiff = self.getLipsVsTeethDiff(pair, lastLips, lastTeeth)
        return angle, candleDiff, lipsVsTeethDiff

    def IsNewBar(self,df, lastCheckedTime, timeFrame):
        new_time = df['time'].iloc[0]
        if lastCheckedTime is None:
            print(f"Первая свеча {timeFrame}, запоминаем время")
            return True, new_time  # Возвращаем флаг новой свечи и новое время
        if new_time != lastCheckedTime:
            print(f"Обнаружена новая свеча {timeFrame}!")
            return True, new_time  # Возвращаем True и новое время
        return False, lastCheckedTime  # Возвращаем False и старое время


    def Df(self,pair, timeFrame):
        bars = mt5.copy_rates_from_pos(pair, timeFrame, 0, 500)
        if bars is None:
            print("Не удалось получить данные:", mt5.last_error())
        df = pd.DataFrame(bars)
        return df

