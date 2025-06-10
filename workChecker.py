import os
from datetime import datetime
from openpyxl import load_workbook
import subprocess

# Загрузка файла Excel
file_path = "alligator_data.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active

# Получаем последнюю дату из столбца A
last_row = sheet.max_row
last_date_str = sheet[f"A{last_row}"].value
workbook.close()

# Проверяем, что значение является строкой с датой
if not last_date_str:
    print("Ошибка: Последняя ячейка пуста или не содержит дату.")
    exit()

# Парсим дату из Excel (предполагаем формат 'YYYY-MM-DD HH:MM:SS')
try:
    last_date = datetime.strptime(str(last_date_str), "%Y-%m-%d %H:%M:%S")
except ValueError:
    print("Ошибка: Неверный формат даты в файле.")
    exit()

# Текущее время
current_date = datetime.now()

# Разница в минутах
time_diff = (current_date - last_date).total_seconds() / 60
print(f"Разница во времени: {time_diff:.2f} минут")

# Если разница >20 минут — перезапускаем процесс
if time_diff > 20:
    print("Разница больше 20 минут. Перезапускаем процесс...")
    # Пример команды для перезапуска процесса (замените на свою)
    process_name = "your_process.exe"  # Например, "notepad.exe"
    subprocess.run(["taskkill", "/f", "/im", process_name], shell=True)
    subprocess.run(["start", process_name], shell=True)
else:
    print("Разница в пределах нормы. Действие не требуется.")